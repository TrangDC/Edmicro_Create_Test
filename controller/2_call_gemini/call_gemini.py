import os
import json
import base64
import requests
from openpyxl import load_workbook
import time
import logging
from enum import Enum
import re

# Thiết lập logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Thay thế bằng API key của bạn
GEMINI_API_KEY = "AIzaSyASlUYT5KMxrtMLVLtUpL7mn4MWOtWf29c"

# Cấu hình Gemini API
generation_config = {
  "temperature": 0.5,
  "top_p": 0.95,
  "top_k": 40,
  "max_output_tokens": 8192,
  "response_mime_type": "text/plain",
}


class SheetType(Enum):
    CHOICE = "TN"
    TRUE_FALSE = "DS"
    SHORT_ANSWER = "TNN"


# Hàm đọc dữ liệu template từ sheet 'prompt_template' của file Excel
def get_prompt_templates_excel(excel_file_path):
    try:
        workbook = load_workbook(excel_file_path)
        if 'prompt_template' not in workbook.sheetnames:
            raise Exception("Không tìm thấy sheet 'prompt_template' trong file Excel.")

        template_sheet = workbook['prompt_template']
        headers = [cell.value for cell in template_sheet[1]]

        prompt_col = headers.index('Prompt')
        subject_col = headers.index('Môn học') if 'Môn học' in headers else None

        # Lấy giá trị prompt duy nhất từ dòng thứ 2
        prompt = template_sheet.cell(row=2, column=prompt_col + 1).value
        subject = template_sheet.cell(row=2, column=subject_col + 1).value if subject_col is not None else None

        if not prompt:
            raise Exception("Không tìm thấy prompt trong sheet 'prompt_template'.")

        return {
            "default": {
                "prompt": prompt,
                "subject": subject if subject else "Mặc định"
            }
        }
    except (KeyError, ValueError) as e:
        raise Exception(f"Lỗi khi đọc template sheet: {e}")
    except Exception as e:
        raise Exception(f"Lỗi khi đọc template sheet: {e}")

# Hàm xử lý một hàng câu hỏi
def process_question_row_gemini(row, col_indices, template_data, gemini_api_key, retries=3):
    question = row[col_indices['question']] if 'question' in col_indices and col_indices['question'] is not None else None
    image_url = row[col_indices['images']] if 'images' in col_indices and col_indices['images'] is not None else None
    correct_answer = row[col_indices['correctAnswer']] if 'correctAnswer' in col_indices and col_indices['correctAnswer'] is not None else 'Không có sẵn đáp án'
    if correct_answer is None:
        correct_answer = 'Không có sẵn đáp án'

    if not question:
        return None

    prompt_text = template_data['prompt'].replace('{question}', str(question)).replace('{correct_answer}',
                                                                                      str(correct_answer)).replace(
        '{môn_học}', str(template_data['subject']))

    gemini_response = None
    json_content = ""
    for attempt in range(retries):
        gemini_response = call_gemini_api(prompt_text, image_url, gemini_api_key, generation_config)
        if gemini_response:
            response = gemini_response[0]
            match = re.search(r'```json\s*(\{.*?\})\s*```', response, re.DOTALL)
            if match:
                json_content = match.group(1)
            logging.info("Gemini response: %s", json_content)
            break
        else:
            logging.info(f"Retry attempt {attempt + 1} failed for question: {question}. Waiting 5 seconds...")
            time.sleep(5)
    if not gemini_response:
        logging.error(f"Failed after {retries} retries for question: {question}")

    return json_content

def get_image_base64(image_url):
    try:
        response = requests.get(image_url, stream=True)
        response.raise_for_status()
        image_bytes = response.content
        return base64.b64encode(image_bytes).decode('utf-8')
    except requests.exceptions.RequestException as e:
        logging.error(f"Lỗi khi tải ảnh từ URL: {image_url}, {e}")
        return None

def call_gemini_api(prompt_text, image_url, gemini_api_key, generation_config):
    headers = {
        "Content-Type": "application/json",
    }

    if image_url:
        try:
            image_base64 = get_image_base64(image_url)
            if image_base64:
                payload = {
                    "contents": [{
                        "parts": [
                            {"text": prompt_text},
                            {
                                "inline_data": {
                                    "mime_type": "image/jpeg",
                                    "data": image_base64
                                }
                            }
                        ]
                    }],
                    "generationConfig": generation_config
                }
                api_url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent?key={gemini_api_key}'
            else:
                api_url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent?key={gemini_api_key}'
                payload = {
                    "contents": [{
                        "parts": [{"text": f"{prompt_text} (Lưu ý: Không thể tải ảnh từ URL: {image_url})" }],
                    }],
                    "generationConfig": generation_config
                }
        except Exception as e:
            api_url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent?key={gemini_api_key}'
            payload = {
                    "contents": [{
                        "parts": [{"text": f"{prompt_text} (Lưu ý: Lỗi khi xử lý ảnh từ URL: {image_url}, {e})" }],
                    }],
                    "generationConfig": generation_config
                }
    else:
        api_url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent?key={gemini_api_key}'
        payload = {
            "contents": [{
                "parts": [{"text": prompt_text}],
            }],
            "generationConfig": generation_config
        }
    try:
        response = requests.post(api_url, headers=headers, json=payload)
        response.raise_for_status()
        response_json = response.json()
        if 'candidates' in response_json and response_json['candidates']:
            return [candidate['content']['parts'][0]['text'] for candidate in response_json['candidates']]
        else:
            logging.error(f"API returned no candidates {response_json}")
            return None
    except requests.exceptions.RequestException as e:
        logging.error(f"Lỗi gọi API: {e}")
        return None
    except KeyError as e:
        logging.error(f"API Error key error: {e}")
        return None
    except json.JSONDecodeError as e:
        logging.error(f"API Error, could not parse response json: {e}")
        return None

def convert_to_json_structure(json_list, sheet_type):
    if sheet_type == SheetType.CHOICE.value:
        return {"choiceQuestions": json_list}
    elif sheet_type == SheetType.TRUE_FALSE.value:
        return {"trueFalseQuestions": json_list}
    elif sheet_type == SheetType.SHORT_ANSWER.value:
        return {"shortQuestions": json_list}
    else:
         return {"unknownQuestions": json_list}

def process_sheet(sheet, templates, sheet_type, all_results):
    logging.info(f"Process sheet gemini {sheet.title}")
    template_data = templates.get("default")
    if not template_data:
        logging.error(f"Không tìm thấy template mặc định")
        return

    headers = [cell.value for cell in sheet[1]]
    col_indices = {
        'question': headers.index('Question') if 'Question' in headers else None,
        'images': headers.index('Images') if 'Images' in headers else None,
        'correctAnswer': headers.index('Correct_Answer') if 'Correct_Answer' in headers else None,
        'gemini': headers.index('Gemini') if 'Gemini' in headers else None
    }
    if col_indices['question'] is None or col_indices['gemini'] is None:
        logging.error(f"Sheet {sheet.title} thiếu các cột bắt buộc Question hoặc Gemini")
        return

    list_json_string = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if row[col_indices['question']]:
            gemini_response = process_question_row_gemini(row, col_indices, template_data, GEMINI_API_KEY)
            if gemini_response:
                try:
                   json_object = json.loads(gemini_response)
                   list_json_string.append(json_object)
                   logging.info(f"Got json data from gemini and added to list")
                except json.JSONDecodeError as e:
                    logging.error(f"Error decoding JSON from Gemini response: {e}")
            time.sleep(1)

    if list_json_string:
       all_results.append(convert_to_json_structure(list_json_string, sheet_type))
        
if __name__ == '__main__':
    excel_file = r"E:\Edmicro\Đề GK2 Toán 10_full lời giải_doc_to_excel_export_20250118_151656.xlsx"
    output_file = r"E:\Edmicro\Tool_tạo_đề\controller\2_call_gemini\output.json"

    templates = get_prompt_templates_excel(excel_file)
    if not templates:
        logging.error("Không có template nào")
        exit()

    workbook = load_workbook(excel_file)
    all_results = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in ['prompt_template', 'tables']:
            continue
        if not (sheet_name.startswith("TN") or sheet_name.startswith("DS") or sheet_name.startswith("TNN")):
            continue

        sheet = workbook[sheet_name]
        sheet_type = sheet_name.split('_')[0]
        try:
             process_sheet(sheet, templates, sheet_type, all_results)
        except Exception as error:
            logging.error(f"Lỗi khi xử lý sheet {sheet_name}: {error}")
            continue
    output_json = {}
    for result in all_results:
        output_json.update(result)
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_json, f, ensure_ascii=False, indent=4)

    logging.info(f"Đã lưu dữ liệu vào file: {output_file}")